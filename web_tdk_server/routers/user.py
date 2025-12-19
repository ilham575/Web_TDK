from fastapi import APIRouter, HTTPException, Depends, status, UploadFile, File
import secrets
from fastapi.security import OAuth2PasswordRequestForm, OAuth2PasswordBearer
from sqlalchemy.orm import Session
from sqlalchemy import and_
from typing import List
from fastapi import Body
from sqlalchemy.exc import IntegrityError
import os
import smtplib
from email.message import EmailMessage
from datetime import timedelta

from schemas.user import User, UserCreate, UserUpdate, Token, ChangePasswordRequest, PasswordResetRequestCreate, PasswordResetRequestResponse, PasswordResetByAdminRequest
from models.user import User as UserModel
from models.password_reset_request import PasswordResetRequest as PasswordResetRequestModel
from database.connection import get_db
from utils.security import hash_password, verify_password, create_access_token, decode_access_token
from typing import List
from sqlalchemy.orm import Session
from io import BytesIO
from fastapi.responses import StreamingResponse

try:
    import openpyxl
    from openpyxl import Workbook
except Exception:
    openpyxl = None

# สร้าง router พร้อมกำหนด prefix
router = APIRouter(prefix="/users", tags=["users"])


def _validate_no_null_classroom_student(db: Session):
    """
    ตรวจสอบใน session ว่ามี ClassroomStudent ที่ถูกสร้าง/แก้ไขโดยยังไม่มี classroom_id หรือไม่
    """
    # import locally to avoid circular imports at module load
    try:
        from models.classroom import ClassroomStudent as _CS
    except Exception:
        _CS = None

    for obj in list(db.new) + list(db.dirty):
        if _CS and isinstance(obj, _CS):
            if getattr(obj, 'classroom_id', None) is None:
                raise HTTPException(status_code=400, detail=f'พบการอัปเดต enrollment ที่ไม่มี classroom_id (id={getattr(obj, "id", None)})')

# กำหนด OAuth2 Security Scheme
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/users/login")

def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    """Get current user from JWT token"""
    try:
        payload = decode_access_token(token)
        username = payload.get("sub")
        if not username:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user = db.query(UserModel).filter(UserModel.username == username).first()
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))


@router.post("/{user_id}/admin_reset")
def admin_reset_user_password(user_id: int, db: Session = Depends(get_db), current_user: UserModel = Depends(get_current_user)):
    """Admin-only: reset a user's password and return a temporary password to the caller.

    NOTE: This endpoint returns the temporary password in the response so the admin can communicate it to the user
    through out-of-band channels (SMS/phone/hand-delivery). Use only in trusted admin contexts.
    """
    if getattr(current_user, 'role', None) != 'admin':
        raise HTTPException(status_code=403, detail='Not authorized')

    user = db.query(UserModel).filter(UserModel.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail='User not found')

    # generate a secure temporary password
    temp_password = secrets.token_urlsafe(8)
    user.hashed_password = hash_password(temp_password)
    user.must_change_password = True
    db.commit()
    # Do NOT log the plaintext password in real systems; return it only to caller here
    return { 'detail': 'password reset', 'temp_password': temp_password }

@router.get("", response_model=List[User])
@router.get("/", response_model=List[User])
def get_users(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    """ดึงข้อมูลผู้ใช้งานทั้งหมด"""
    users = db.query(UserModel).offset(skip).limit(limit).all()
    return users

@router.post("", response_model=User, status_code=status.HTTP_201_CREATED)
@router.post("/", response_model=User, status_code=status.HTTP_201_CREATED)
def create_user(user: UserCreate, db: Session = Depends(get_db)):
    """เพิ่มผู้ใช้งานใหม่"""
    # Check if username already exists
    db_user = db.query(UserModel).filter(UserModel.username == user.username).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check if email already exists
    db_email = db.query(UserModel).filter(UserModel.email == user.email).first()
    if db_email:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    # Hash password and create user
    hashed_password = hash_password(user.password)
    db_user = UserModel(
        username=user.username,
        email=user.email,
        full_name=user.full_name,
        hashed_password=hashed_password,
        role=user.role,
        school_id=user.school_id,
        grade_level=user.grade_level
    )
    
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

@router.post("/login", response_model=Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    """ล็อกอินและสร้าง JWT"""
    user = db.query(UserModel).filter(UserModel.username == form_data.username).first()
    
    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid username or password"
        )
    
    if not user.is_active:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="User account is inactive"
        )
    
    access_token = create_access_token(data={"sub": user.username})
    return {
        "access_token": access_token, 
        "token_type": "bearer",
        "user_info": user
    }

@router.get("/me", response_model=User)
def get_current_user_info(current_user: UserModel = Depends(get_current_user)):
    """ดึงข้อมูลผู้ใช้งานปัจจุบันจาก JWT"""
    return current_user


@router.post('/forgot_password')
def forgot_password(email: str = Body(..., embed=True), db: Session = Depends(get_db)):
    """Initiate password reset: generate a short-lived token and email it to user (if configured).

    Returns a generic success message. If SMTP is not configured and RETURN_RESET_TOKEN=true, the token
    will be returned in the response (useful for local dev).
    """
    user = db.query(UserModel).filter(UserModel.email == email).first()
    # always respond with generic message to avoid user enumeration
    generic = {"detail": "If an account with that email exists, a reset link has been sent."}
    if not user:
        return generic

    # create reset token (1 hour validity)
    token = create_access_token({"sub": user.username, "action": "reset_password"}, expires_delta=timedelta(hours=1))

    # Attempt to send email if SMTP configured
    smtp_host = os.getenv('SMTP_HOST')
    smtp_port = os.getenv('SMTP_PORT')
    smtp_user = os.getenv('SMTP_USER')
    smtp_pass = os.getenv('SMTP_PASS')
    from_addr = os.getenv('EMAIL_FROM') or smtp_user
    frontend_url = os.getenv('FRONTEND_URL', 'http://localhost:3000')

    if smtp_host and smtp_user and smtp_pass:
        try:
            reset_link = f"{frontend_url}/reset-password?token={token}"
            msg = EmailMessage()
            msg['Subject'] = 'Password reset request'
            msg['From'] = from_addr
            msg['To'] = email
            msg.set_content(f'You requested a password reset. Click the link to reset: {reset_link}\nIf you did not request this, ignore this email.')

            with smtplib.SMTP(smtp_host, int(smtp_port or 25)) as smtp:
                smtp.starttls()
                smtp.login(smtp_user, smtp_pass)
                smtp.send_message(msg)
        except Exception as e:
            # log and still return generic
            print('failed to send reset email', e)
            return generic
        return generic

    # If SMTP not configured, optionally return token for dev convenience
    if os.getenv('RETURN_RESET_TOKEN', 'false').lower() == 'true':
        return {"detail": "No SMTP configured; returning token for development.", "reset_token": token}

    return generic


@router.post('/request_password_reset')
def request_password_reset(request_data: PasswordResetRequestCreate, db: Session = Depends(get_db)):
    """
    Request password reset by username.
    - Teacher/Student: Request goes to admin of their school
    - Admin: Request goes to owner
    """
    username = request_data.username
    user = db.query(UserModel).filter(UserModel.username == username).first()
    
    # Generic response to avoid username enumeration
    generic_success = {"detail": "หากชื่อผู้ใช้ถูกต้อง คำขอรีเซ็ตรหัสผ่านจะถูกส่งไปยังผู้ดูแลระบบ", "success": True}
    
    if not user:
        return generic_success
    
    # Check if there's already a pending request
    existing_request = db.query(PasswordResetRequestModel).filter(
        PasswordResetRequestModel.user_id == user.id,
        PasswordResetRequestModel.status == "pending"
    ).first()
    
    if existing_request:
        return {"detail": "คุณมีคำขอรีเซ็ตรหัสผ่านที่รอดำเนินการอยู่แล้ว กรุณารอผู้ดูแลระบบอนุมัติ", "success": True}
    
    # Create password reset request
    reset_request = PasswordResetRequestModel(
        user_id=user.id,
        username=user.username,
        full_name=user.full_name,
        email=user.email,
        role=user.role,
        school_id=user.school_id,
        status="pending"
    )
    
    db.add(reset_request)
    db.commit()
    db.refresh(reset_request)
    
    # Determine who should handle this request
    if user.role in ['teacher', 'student']:
        return {
            "detail": "คำขอรีเซ็ตรหัสผ่านถูกส่งไปยังแอดมินของโรงเรียนแล้ว กรุณารอการอนุมัติ",
            "success": True,
            "target": "admin"
        }
    elif user.role == 'admin':
        return {
            "detail": "คำขอรีเซ็ตรหัสผ่านถูกส่งไปยัง Owner แล้ว กรุณารอการอนุมัติ",
            "success": True,
            "target": "owner"
        }
    
    return generic_success


@router.get('/password_reset_requests', response_model=List[PasswordResetRequestResponse])
def get_password_reset_requests(
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    """
    Get pending password reset requests.
    - Admin: sees requests from teacher/student in their school
    - Owner: sees requests from admins
    """
    role = getattr(current_user, 'role', None)
    
    if role == 'admin':
        # Admin sees requests from teachers/students in their school
        requests = db.query(PasswordResetRequestModel).filter(
            PasswordResetRequestModel.school_id == current_user.school_id,
            PasswordResetRequestModel.role.in_(['teacher', 'student']),
            PasswordResetRequestModel.status == "pending"
        ).order_by(PasswordResetRequestModel.created_at.desc()).all()
    elif role == 'owner':
        # Owner sees requests from admins
        requests = db.query(PasswordResetRequestModel).filter(
            PasswordResetRequestModel.role == 'admin',
            PasswordResetRequestModel.status == "pending"
        ).order_by(PasswordResetRequestModel.created_at.desc()).all()
    else:
        raise HTTPException(status_code=403, detail="ไม่มีสิทธิ์เข้าถึง")
    
    return requests


@router.post('/password_reset_requests/{request_id}/approve')
def approve_password_reset(
    request_id: int,
    reset_data: PasswordResetByAdminRequest,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    """
    Approve a password reset request and set new password.
    - Admin: can approve requests from teacher/student in their school
    - Owner: can approve requests from admins
    """
    role = getattr(current_user, 'role', None)
    
    reset_request = db.query(PasswordResetRequestModel).filter(
        PasswordResetRequestModel.id == request_id
    ).first()
    
    if not reset_request:
        raise HTTPException(status_code=404, detail="ไม่พบคำขอรีเซ็ตรหัสผ่าน")
    
    if reset_request.status != "pending":
        raise HTTPException(status_code=400, detail="คำขอนี้ถูกดำเนินการไปแล้ว")
    
    # Check authorization
    if role == 'admin':
        if reset_request.role not in ['teacher', 'student']:
            raise HTTPException(status_code=403, detail="ไม่มีสิทธิ์อนุมัติคำขอนี้")
        if reset_request.school_id != current_user.school_id:
            raise HTTPException(status_code=403, detail="ไม่มีสิทธิ์อนุมัติคำขอจากโรงเรียนอื่น")
    elif role == 'owner':
        if reset_request.role != 'admin':
            raise HTTPException(status_code=403, detail="ไม่มีสิทธิ์อนุมัติคำขอนี้")
    else:
        raise HTTPException(status_code=403, detail="ไม่มีสิทธิ์อนุมัติคำขอ")
    
    # Find the user and reset their password
    user = db.query(UserModel).filter(UserModel.id == reset_data.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="ไม่พบผู้ใช้")
    
    # Update password
    user.hashed_password = hash_password(reset_data.new_password)
    user.must_change_password = True
    
    # Update request status
    reset_request.status = "approved"
    
    db.commit()
    
    return {
        "detail": f"รีเซ็ตรหัสผ่านสำหรับ {user.username} เรียบร้อยแล้ว",
        "success": True,
        "username": user.username
    }


@router.post('/password_reset_requests/{request_id}/reject')
def reject_password_reset(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    """
    Reject a password reset request.
    """
    role = getattr(current_user, 'role', None)
    
    reset_request = db.query(PasswordResetRequestModel).filter(
        PasswordResetRequestModel.id == request_id
    ).first()
    
    if not reset_request:
        raise HTTPException(status_code=404, detail="ไม่พบคำขอรีเซ็ตรหัสผ่าน")
    
    if reset_request.status != "pending":
        raise HTTPException(status_code=400, detail="คำขอนี้ถูกดำเนินการไปแล้ว")
    
    # Check authorization
    if role == 'admin':
        if reset_request.role not in ['teacher', 'student']:
            raise HTTPException(status_code=403, detail="ไม่มีสิทธิ์ปฏิเสธคำขอนี้")
        if reset_request.school_id != current_user.school_id:
            raise HTTPException(status_code=403, detail="ไม่มีสิทธิ์ปฏิเสธคำขอจากโรงเรียนอื่น")
    elif role == 'owner':
        if reset_request.role != 'admin':
            raise HTTPException(status_code=403, detail="ไม่มีสิทธิ์ปฏิเสธคำขอนี้")
    else:
        raise HTTPException(status_code=403, detail="ไม่มีสิทธิ์ปฏิเสธคำขอ")
    
    # Update request status
    reset_request.status = "rejected"
    db.commit()
    
    return {
        "detail": "ปฏิเสธคำขอรีเซ็ตรหัสผ่านเรียบร้อยแล้ว",
        "success": True
    }


@router.post('/reset_password')
def reset_password(token: str = Body(...), new_password: str = Body(...), db: Session = Depends(get_db)):
    """Reset a user's password using a valid reset token."""
    try:
        payload = decode_access_token(token)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if payload.get('action') != 'reset_password':
        raise HTTPException(status_code=400, detail='Invalid token action')

    username = payload.get('sub')
    if not username:
        raise HTTPException(status_code=400, detail='Invalid token payload')

    user = db.query(UserModel).filter(UserModel.username == username).first()
    if not user:
        raise HTTPException(status_code=404, detail='User not found')

    # update password
    user.hashed_password = hash_password(new_password)
    db.commit()
    return {"detail": "Password updated successfully"}


@router.post('/bulk_upload')
def bulk_upload_users(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: UserModel = Depends(get_current_user)):
    """Upload an Excel (.xlsx) file to bulk-create users.

    Expected columns (first row header): username,email,full_name,password,role,school_id (optional)
    Only admins are allowed to use this endpoint. If school_id is missing for a row, current_user.school_id is used.
    Returns a summary of created users and per-row errors.
    """
    if getattr(current_user, 'role', None) != 'admin':
        raise HTTPException(status_code=403, detail='Not authorized')

    if openpyxl is None:
        raise HTTPException(status_code=500, detail='Server missing openpyxl dependency')

    content = file.file.read()
    try:
        # data_only=True reads calculated formula values instead of formula strings
        wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'Failed to read Excel file: {str(e)}')

    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail='Excel file must contain a header row and at least one data row')

    header = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
    # map header names to indexes
    idx = {name: i for i, name in enumerate(header)}
    required_cols = ['username', 'email', 'full_name', 'password', 'role']
    for col in required_cols:
        if col not in idx:
            raise HTTPException(status_code=400, detail=f'Missing required column: {col}')

    created = []
    errors = []
    for r_i, row in enumerate(rows[1:], start=2):
        try:
            username = str(row[idx['username']]).strip() if row[idx['username']] is not None else ''
            email = str(row[idx['email']]).strip() if row[idx['email']] is not None else ''
            full_name = str(row[idx['full_name']]).strip() if row[idx['full_name']] is not None else ''
            password = str(row[idx['password']]).strip() if row[idx['password']] is not None else ''
            role = str(row[idx['role']]).strip() if row[idx['role']] is not None else ''
            school_id = None
            if 'school_id' in idx and row[idx['school_id']] is not None:
                try:
                    school_id = int(row[idx['school_id']])
                except Exception:
                    school_id = None
            
            grade_level = None
            if 'grade_level' in idx and row[idx['grade_level']] is not None:
                grade_level = str(row[idx['grade_level']]).strip()

            if not username or not email or not password or role not in ('teacher', 'student'):
                raise ValueError('Invalid data - required fields missing or role invalid')

            # default school_id to admin's school if not provided
            if school_id is None:
                school_id = getattr(current_user, 'school_id', None)

            # uniqueness checks
            existing_user = db.query(UserModel).filter((UserModel.username == username) | (UserModel.email == email)).first()
            if existing_user:
                raise ValueError('Username or email already exists')

            hashed = hash_password(password)
            new_user = UserModel(username=username, email=email, full_name=full_name, hashed_password=hashed, role=role, school_id=school_id, grade_level=grade_level)
            db.add(new_user)
            db.flush()
            created.append({'row': r_i, 'username': username, 'id': new_user.id})
        except Exception as e:
            errors.append({'row': r_i, 'error': str(e)})

    # commit only after processing all rows to keep atomic-ish behavior for created set
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f'Failed to commit created users: {str(e)}')

    return { 'created_count': len(created), 'created': created, 'errors': errors }

@router.put("/me", response_model=User)
def update_current_user(
    user_update: UserUpdate,
    current_user: UserModel = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """อัปเดตข้อมูลผู้ใช้งานปัจจุบัน"""
    # authorization checks
    incoming = user_update.dict(exclude_unset=True)

    # Only admins can change role, is_active or school_id
    if 'role' in incoming and getattr(current_user, 'role', None) != 'admin':
        raise HTTPException(status_code=403, detail='Only admins can change user role')
    if 'is_active' in incoming and getattr(current_user, 'role', None) != 'admin':
        raise HTTPException(status_code=403, detail='Only admins can change account activation')
    if 'school_id' in incoming and getattr(current_user, 'role', None) != 'admin':
        raise HTTPException(status_code=403, detail='Only admins can change school assignment')

    # Only admins or teachers can update grade_level (prevent students from modifying their own class)
    if 'grade_level' in incoming and getattr(current_user, 'role', None) not in ('admin', 'teacher'):
        raise HTTPException(status_code=403, detail='Only admin or teacher can update grade_level')

    for field, value in incoming.items():
        setattr(current_user, field, value)
    
    db.commit()
    db.refresh(current_user)
    return current_user

@router.post("/change_password")
def change_password(
    password_data: ChangePasswordRequest,
    current_user: UserModel = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """เปลี่ยนรหัสผ่าน"""
    # If the account is not in forced-change state, require current password verification
    if not getattr(current_user, 'must_change_password', False):
        if not password_data.current_password or not verify_password(password_data.current_password, current_user.hashed_password):
            raise HTTPException(status_code=400, detail="รหัสผ่านปัจจุบันไม่ถูกต้อง")

    # Update to new password (allowed when forced change or after successful verification)
    current_user.hashed_password = hash_password(password_data.new_password)
    current_user.must_change_password = False  # Clear the flag
    db.commit()

    return {"message": "เปลี่ยนรหัสผ่านเรียบร้อยแล้ว"}


@router.get("/{user_id}/deletion_status")
def check_user_deletion_status(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    """Admin-only: Check if a user can be deleted and why/why not"""
    if getattr(current_user, 'role', None) != 'admin':
        raise HTTPException(status_code=403, detail='Only admins can check deletion status')
    
    user = db.query(UserModel).filter(UserModel.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail='User not found')
    
    # Check all conditions
    can_delete = True
    reasons = []
    
    # Check if user is active
    if user.is_active:
        can_delete = False
        reasons.append('ผู้ใช้ยังใช้งานได้อยู่ - ต้องปิดใช้งานก่อน')
    
    # Check if user is another admin
    if user.role == 'admin':
        can_delete = False
        reasons.append('ไม่สามารถลบแอดมินคนอื่นได้')
    
    # Check if it's the current admin
    if current_user.id == user_id:
        can_delete = False
        reasons.append('ไม่สามารถลบบัญชีตัวเองได้')
    
    # For teachers, check active subjects
    if user.role == 'teacher':
        try:
            from models.subject import Subject as SubjectModel
            active_subjects = db.query(SubjectModel).filter(
                SubjectModel.teacher_id == user_id,
                SubjectModel.is_ended == False
            ).count()
            if active_subjects > 0:
                can_delete = False
                reasons.append(f'ครูมีรายวิชาที่ยังใช้งานอยู่ {active_subjects} รายวิชา - ต้องจบคอร์สก่อน')
        except ImportError:
            pass
    
    # For students, check enrollments in active subjects only
    if user.role == 'student':
        try:
            from models.subject_student import SubjectStudent as SubjectStudentModel
            from models.subject import Subject as SubjectModel
            active_enrollments = db.query(SubjectStudentModel).join(
                SubjectModel, SubjectStudentModel.subject_id == SubjectModel.id
            ).filter(
                SubjectStudentModel.student_id == user_id,
                SubjectModel.is_ended == False
            ).count()
            if active_enrollments > 0:
                can_delete = False
                reasons.append(f'นักเรียนยังลงทะเบียนรายวิชาที่กำลังเรียนอยู่ {active_enrollments} รายวิชา - ต้องจบคอร์สหรือถอนออกก่อน')
        except ImportError:
            pass
    
    return {
        'user_id': user_id,
        'username': user.username,
        'role': user.role,
        'is_active': user.is_active,
        'can_delete': can_delete,
        'reasons': reasons,
        'next_steps': [] if can_delete else [
            'ปิดใช้งานผู้ใช้' if user.is_active else None,
            'จบคอร์สทั้งหมด' if user.role == 'teacher' and not can_delete else None,
            'ถอนนักเรียนออกจากรายวิชา' if user.role == 'student' and not can_delete else None
        ]
    }


@router.delete("/{user_id}")
def delete_user(
    user_id: int, 
    db: Session = Depends(get_db), 
    current_user: UserModel = Depends(get_current_user)
):
    """Admin-only: Delete a user with safety checks"""
    # Only admins can delete users
    if getattr(current_user, 'role', None) != 'admin':
        raise HTTPException(status_code=403, detail='Only admins can delete users')
    
    # Cannot delete yourself
    if current_user.id == user_id:
        raise HTTPException(status_code=400, detail='Cannot delete your own account')
    
    # Find the user to delete
    user_to_delete = db.query(UserModel).filter(UserModel.id == user_id).first()
    if not user_to_delete:
        raise HTTPException(status_code=404, detail='User not found')
    
    # Safety checks before deletion
    deletion_blocks = []
    
    # Check if user is active
    if user_to_delete.is_active:
        deletion_blocks.append('User is still active. Deactivate the user first before deletion.')
    
    # Check if user is another admin (prevent deleting other admins)
    if user_to_delete.role == 'admin':
        deletion_blocks.append('Cannot delete another admin account.')
    
    # For teachers, check if they have active subjects
    if user_to_delete.role == 'teacher':
        try:
            from models.subject import Subject as SubjectModel
            active_subjects = db.query(SubjectModel).filter(
                SubjectModel.teacher_id == user_id,
                SubjectModel.is_ended == False
            ).count()
            if active_subjects > 0:
                deletion_blocks.append(f'Teacher has {active_subjects} active subject(s). End all subjects first.')
        except ImportError:
            # If Subject model doesn't exist, skip this check
            pass
    
    # For students, check if they have active enrollments in non-ended subjects
    if user_to_delete.role == 'student':
        try:
            from models.subject_student import SubjectStudent as SubjectStudentModel
            from models.subject import Subject as SubjectModel
            active_enrollments = db.query(SubjectStudentModel).join(
                SubjectModel, SubjectStudentModel.subject_id == SubjectModel.id
            ).filter(
                SubjectStudentModel.student_id == user_id,
                SubjectModel.is_ended == False
            ).count()
            if active_enrollments > 0:
                deletion_blocks.append(f'Student is enrolled in {active_enrollments} active subject(s). End subjects or remove enrollments first.')
        except ImportError:
            # If models don't exist, skip this check
            pass
    
    # If there are any blocks, return them as errors
    if deletion_blocks:
        raise HTTPException(
            status_code=400, 
            detail={
                'message': 'Cannot delete user due to the following reasons:',
                'blocks': deletion_blocks
            }
        )
    
    # If all checks pass, perform deletion
    try:
        db.delete(user_to_delete)
        db.commit()
        return {'message': f'User {user_to_delete.username} deleted successfully'}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f'Failed to delete user: {str(e)}')


@router.patch("/{user_id}/grade_level")
def admin_update_student_grade_level(
    user_id: int,
    grade_level: str = Body(..., embed=True),
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    """Admin/Teacher: Update a student's grade level

    - Admin: can update any student within their school
    - Teacher: can update students in their school (limited to students only)
    """
    role = getattr(current_user, 'role', None)
    if role not in ('admin', 'teacher'):
        raise HTTPException(status_code=403, detail='Only admins or teachers can update grade levels')
    
    # Check if user exists
    user = db.query(UserModel).filter(UserModel.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail='ไม่พบผู้ใช้ที่ระบุ')
    
    # Allow updating grade level for students (or bulk create new students)
    if user.role != 'student':
        raise HTTPException(status_code=400, detail='เฉพาะนักเรียนเท่านั้นที่สามารถเพิ่มชั้นเรียนได้')

    # If teacher is performing the update, ensure same school
    if role == 'teacher' and user.school_id != current_user.school_id:
        raise HTTPException(status_code=403, detail='ครูไม่มีสิทธิ์แก้ไขชั้นเรียนของนักเรียนจากโรงเรียนอื่น')
    
    # Update grade level
    user.grade_level = grade_level
    db.commit()
    db.refresh(user)
    
    return {'message': f'อัปเดตชั้นเรียนของ {user.full_name} เป็น {grade_level} เรียบร้อยแล้ว', 'user': user}


@router.post("/bulk_assign_grade")
def bulk_assign_grade_to_students(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    """Admin-only: Bulk assign grade levels to existing or new students via Excel.
    
    Expected columns: username, email, full_name, grade_level
    If student exists (by username or email), update their grade level.
    If student doesn't exist, create a new student with temporary password.
    """
    if getattr(current_user, 'role', None) != 'admin':
        raise HTTPException(status_code=403, detail='Not authorized')

    if openpyxl is None:
        raise HTTPException(status_code=500, detail='Server missing openpyxl dependency')

    content = file.file.read()
    try:
        # data_only=True reads calculated formula values instead of formula strings
        wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'Failed to read Excel file: {str(e)}')

    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail='Excel file must contain a header row and at least one data row')

    header = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    
    required_cols = ['username', 'email', 'full_name', 'grade_level']
    for col in required_cols:
        if col not in idx:
            raise HTTPException(status_code=400, detail=f'Missing required column: {col}')

    updated = []
    created = []
    errors = []
    
    for r_i, row in enumerate(rows[1:], start=2):
        try:
            username = row[idx['username']].strip() if row[idx['username']] is not None else ''
            email = row[idx['email']].strip() if row[idx['email']] is not None else ''
            full_name = row[idx['full_name']].strip() if row[idx['full_name']] is not None else ''
            grade_level = row[idx['grade_level']].strip() if row[idx['grade_level']] is not None else ''

            if not username or not email or not full_name or not grade_level:
                raise ValueError('Invalid data - required fields missing')

            # Try to find existing student by username or email
            existing_student = db.query(UserModel).filter(
                (UserModel.username == username) | (UserModel.email == email),
                UserModel.role == 'student'
            ).first()

            if existing_student:
                # Update existing student's grade level
                existing_student.grade_level = grade_level
                db.flush()
                updated.append({
                    'row': r_i,
                    'username': username,
                    'id': existing_student.id,
                    'action': 'updated',
                    'grade_level': grade_level
                })
            else:
                # Create new student with temporary password
                temp_password = secrets.token_urlsafe(8)
                hashed_pass = hash_password(temp_password)
                
                new_student = UserModel(
                    username=username,
                    email=email,
                    full_name=full_name,
                    hashed_password=hashed_pass,
                    role='student',
                    school_id=getattr(current_user, 'school_id', None),
                    grade_level=grade_level,
                    must_change_password=True
                )
                db.add(new_student)
                db.flush()
                created.append({
                    'row': r_i,
                    'username': username,
                    'id': new_student.id,
                    'action': 'created',
                    'grade_level': grade_level,
                    'temp_password': temp_password
                })
        except Exception as e:
            errors.append({'row': r_i, 'error': str(e)})

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f'Failed to commit: {str(e)}')

    return {
        'updated_count': len(updated),
        'created_count': len(created),
        'error_count': len(errors),
        'updated': updated,
        'created': created,
        'errors': errors,
        'message': f'อัปเดต {len(updated)} นักเรียน สร้างใหม่ {len(created)} นักเรียน'
    }


@router.patch("/{user_id}/deactivate")
def deactivate_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    """Admin-only: Deactivate a user (set is_active to False)"""
    if getattr(current_user, 'role', None) != 'admin':
        raise HTTPException(status_code=403, detail='Only admins can deactivate users')
    
    if current_user.id == user_id:
        raise HTTPException(status_code=400, detail='Cannot deactivate your own account')
    
    user = db.query(UserModel).filter(UserModel.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail='User not found')
    
    if user.role == 'admin':
        raise HTTPException(status_code=400, detail='Cannot deactivate another admin account')
    
    user.is_active = False
    db.commit()
    db.refresh(user)
    return {'message': f'User {user.username} deactivated successfully', 'user': user}


@router.patch("/{user_id}/activate")
def activate_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    """Admin-only: Activate a user (set is_active to True)"""
    if getattr(current_user, 'role', None) != 'admin':
        raise HTTPException(status_code=403, detail='Only admins can activate users')
    
    user = db.query(UserModel).filter(UserModel.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail='User not found')
    
    user.is_active = True
    db.commit()
    db.refresh(user)
    return {'message': f'User {user.username} activated successfully', 'user': user}


@router.get('/bulk_template')
def download_bulk_template(current_user: UserModel = Depends(get_current_user)):
    """Return an Excel template (.xlsx) for admins to fill and upload via /users/bulk_upload."""
    if getattr(current_user, 'role', None) != 'admin':
        raise HTTPException(status_code=403, detail='Not authorized')

    if openpyxl is None:
        raise HTTPException(status_code=500, detail='Server missing openpyxl dependency')

    wb = Workbook()
    ws = wb.active
    # Header row
    headers = ['username', 'email', 'full_name', 'password', 'role', 'school_id', 'grade_level']
    ws.append(headers)
    # Example rows
    ws.append(['alice', 'alice@example.com', 'Alice A', 'secret123', 'teacher', '', ''])
    ws.append(['bob', 'bob@example.com', 'Bob B', 'p@ssw0rd', 'student', '', 'ป.1'])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = 'user_bulk_template.xlsx'
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    return StreamingResponse(stream, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)


@router.post("/promote_students")
def promote_students(
    promotion_type: str = Body(..., embed=True),  # "mid_term", "mid_term_with_promotion" or "end_of_year"
    student_ids: List[int] = Body(..., embed=True),
    new_grade_level: str = Body(None, embed=True),  # Required for mid_term_with_promotion and end_of_year
    new_academic_year: str = Body(None, embed=True),  # For end_of_year
    classroom_id: int = Body(None, embed=True),
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    """
    Promote individual students
    
    promotion_type: 
      - "mid_term": move to next semester (same academic year)
      - "mid_term_with_promotion": move to next semester + new grade level (same academic year)
      - "end_of_year": move to new grade level (new academic year)
    
    new_grade_level: Required for mid_term_with_promotion and end_of_year promotion
    new_academic_year: For end_of_year, auto +1 if not provided
    """
    from models.classroom import ClassroomStudent, Classroom
    
    # Only admins can promote students
    if getattr(current_user, 'role', None) != 'admin':
        raise HTTPException(status_code=403, detail='Only admins can promote students')
    
    if not student_ids:
        raise HTTPException(status_code=400, detail='No students selected')
    
    if promotion_type not in ['mid_term', 'mid_term_with_promotion', 'end_of_year']:
        raise HTTPException(
            status_code=400, 
            detail='Invalid promotion_type. Must be "mid_term", "mid_term_with_promotion", or "end_of_year"'
        )
    
    if (promotion_type in ['mid_term_with_promotion', 'end_of_year']) and not new_grade_level:
        raise HTTPException(
            status_code=400, 
            detail=f'new_grade_level is required for {promotion_type} promotion'
        )
    
    promoted_count = 0
    failed_count = 0
    errors = []
    promoted_students = []
    source_classroom = None
    target_classroom = None
    
    try:
        if classroom_id is not None:
            source_classroom = db.query(Classroom).filter(Classroom.id == classroom_id).first()
            if not source_classroom:
                raise HTTPException(status_code=404, detail=f'ไม่พบข้อมูลชั้นเรียน ID {classroom_id}')

        def compose_class_name(grade_level_value, room_value):
            return grade_level_value if not room_value else f"{grade_level_value}/{room_value}"

        def ensure_target_classroom():
            nonlocal target_classroom
            if target_classroom or not source_classroom:
                return target_classroom

            # Determine target classroom attributes
            if promotion_type == 'mid_term':
                new_semester = 2
                new_academic_year_value = source_classroom.academic_year
                new_grade = source_classroom.grade_level
                # For mid_term, keep the same classroom name
                new_name = source_classroom.name
            elif promotion_type == 'mid_term_with_promotion':
                new_semester = 2
                new_academic_year_value = source_classroom.academic_year
                new_grade = new_grade_level
                # Generate new classroom name based on selected grade level
                new_name = compose_class_name(new_grade, source_classroom.room_number)
                # Try to find existing classroom with matching name pattern for new grade
                potential_target = db.query(Classroom).filter(
                    and_(
                        Classroom.grade_level == new_grade,
                        Classroom.school_id == source_classroom.school_id,
                        Classroom.semester == new_semester,
                        Classroom.academic_year == new_academic_year_value,
                        Classroom.is_active == True
                    )
                ).first()
                if potential_target:
                    new_name = potential_target.name
            else:  # end_of_year
                new_semester = 1
                new_academic_year_value = new_academic_year or str(int(source_classroom.academic_year or '0') + 1)
                new_grade = new_grade_level
                # Generate new classroom name based on selected grade level
                new_name = compose_class_name(new_grade, source_classroom.room_number)
                # Try to find existing classroom with matching name pattern for new grade
                potential_target = db.query(Classroom).filter(
                    and_(
                        Classroom.grade_level == new_grade,
                        Classroom.school_id == source_classroom.school_id,
                        Classroom.semester == new_semester,
                        Classroom.academic_year == new_academic_year_value,
                        Classroom.is_active == True
                    )
                ).first()
                if potential_target:
                    new_name = potential_target.name

            room_number = source_classroom.room_number

            existing = db.query(Classroom).filter(
                and_(
                    Classroom.name == new_name,
                    Classroom.school_id == source_classroom.school_id,
                    Classroom.semester == new_semester,
                    Classroom.academic_year == new_academic_year_value,
                    Classroom.is_active == True
                )
            ).first()

            if existing:
                target_classroom = existing
                return target_classroom

            target_classroom = Classroom(
                name=new_name,
                grade_level=new_grade,
                room_number=room_number,
                semester=new_semester,
                academic_year=new_academic_year_value,
                school_id=source_classroom.school_id,
                parent_classroom_id=source_classroom.id
            )
            db.add(target_classroom)
            db.flush()
            return target_classroom

        def deactivate_old_enrollment(enrollment_id_value):
            if enrollment_id_value:
                db.query(ClassroomStudent).filter(
                    ClassroomStudent.id == enrollment_id_value
                ).update(
                    {ClassroomStudent.is_active: False},
                    synchronize_session=False
                )

        for student_id in student_ids:
            try:
                student = db.query(UserModel).filter(
                    UserModel.id == student_id,
                    UserModel.role == 'student'
                ).first()

                if not student:
                    # ลองค้นหาโดยไม่ตรวจสอบ role ก่อน เพื่อดูว่าพบ user ไหม
                    user_exists = db.query(UserModel).filter(UserModel.id == student_id).first()
                    if user_exists:
                        errors.append(f'⚠️ ID {student_id} เป็น {user_exists.role} ไม่ใช่นักเรียน')
                    else:
                        errors.append(f'⚠️ ไม่พบนักเรียน ID {student_id}')
                    failed_count += 1
                    continue
                
                old_grade = student.grade_level or 'ไม่ระบุ'
                
                # ดึงข้อมูลชั้นเรียนปัจจุบันของนักเรียน - ดึงแค่ ID และข้อมูลที่จำเป็น
                enrollment_data = db.query(
                    ClassroomStudent.id,
                    ClassroomStudent.classroom_id,
                    Classroom.semester
                ).join(
                    Classroom, ClassroomStudent.classroom_id == Classroom.id
                ).filter(
                    ClassroomStudent.student_id == student_id,
                    ClassroomStudent.is_active == True
                ).first()
                
                if promotion_type == 'mid_term':
                    # ต้องมีข้อมูลชั้นเรียนเพื่อเลื่อนเทอม
                    if not enrollment_data:
                        failed_count += 1
                        errors.append(f'⚠️ นักเรียน ID {student_id} ไม่อยู่ในชั้นเรียนใด')
                        continue

                    # enrollment_data is (enrollment_id, classroom_id, semester)
                    enrollment_id, classroom_id, current_semester = enrollment_data
                    if current_semester == 2:
                        failed_count += 1
                        errors.append(f'⚠️ นักเรียน ID {student_id} อยู่เทอม 2 แล้ว')
                        continue
                    
                    # อัพเดต grade_level: เทอม 1 → เทอม 2
                    if 'เทอม 2' not in str(student.grade_level or ''):
                        if 'เทอม 1' in str(student.grade_level or ''):
                            student.grade_level = str(student.grade_level).replace('เทอม 1', 'เทอม 2')
                        else:
                            student.grade_level = f"{student.grade_level} (เทอม 2)" if student.grade_level else "เทอม 2"
                    deactivate_old_enrollment(enrollment_id)
                    
                elif promotion_type in ['mid_term_with_promotion', 'end_of_year']:
                    # ต้องมีข้อมูลชั้นเรียนเพื่อเลื่อน
                    if not enrollment_data:
                        failed_count += 1
                        errors.append(f'⚠️ นักเรียน ID {student_id} ไม่อยู่ในชั้นเรียนใด')
                        continue
                    
                    # enrollment_data is (enrollment_id, classroom_id, semester)
                    enrollment_id, classroom_id, current_semester = enrollment_data
                    
                    if promotion_type == 'mid_term_with_promotion':
                        if current_semester == 2:
                            failed_count += 1
                            errors.append(f'⚠️ นักเรียน ID {student_id} อยู่เทอม 2 แล้ว')
                            continue
                    elif promotion_type == 'end_of_year':
                        pass
                    
                    # อัพเดต grade_level
                    student.grade_level = new_grade_level
                    deactivate_old_enrollment(enrollment_id)
                
                target = ensure_target_classroom()
                if target:
                    existing_target = db.query(ClassroomStudent).filter(
                        ClassroomStudent.classroom_id == target.id,
                        ClassroomStudent.student_id == student_id
                    ).first()
                    if existing_target:
                        if not existing_target.is_active:
                            existing_target.is_active = True
                    else:
                        db.add(ClassroomStudent(
                            classroom_id=target.id,
                            student_id=student_id
                        ))

                db.add(student)
                db.flush()
                promoted_count += 1
                promoted_students.append({
                    'id': student.id,
                    'username': student.username,
                    'full_name': student.full_name,
                    'old_grade': old_grade,
                    'new_grade': student.grade_level
                })
                
            except Exception as e:
                failed_count += 1
                errors.append(f'⚠️ นักเรียน ID {student_id}: {str(e)}')
        
        # Validate session objects and commit all changes
        _validate_no_null_classroom_student(db)
        
        try:
            db.commit()
        except IntegrityError as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f'Database integrity error during student promotion: {str(e)}')
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f'Failed to promote students: {str(e)}')
    
    return {
        'success': True,
        'promotion_type': promotion_type,
        'promoted_count': promoted_count,
        'failed_count': failed_count,
        'promoted_students': promoted_students,
        'errors': errors,
        'message': f'✓ เลื่อนชั้นเรียนสำเร็จ {promoted_count} นักเรียน' + (f', ล้มเหลว {failed_count} นักเรียน' if failed_count > 0 else '')
    }


@router.post("/promote_template")
def download_promote_template(
    promotion_type: str = Body(..., embed=True),  # "mid_term" or "end_of_year"
    current_user: UserModel = Depends(get_current_user)
):
    """Generate and download Excel template for promoting students"""
    if getattr(current_user, 'role', None) != 'admin':
        raise HTTPException(status_code=403, detail='Not authorized')
    
    if openpyxl is None:
        raise HTTPException(status_code=500, detail='Server missing openpyxl dependency')
    
    if promotion_type not in ['mid_term', 'end_of_year']:
        raise HTTPException(status_code=400, detail='Invalid promotion_type')
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Promote'
    
    if promotion_type == 'mid_term':
        ws['A1'] = 'คำแนะนำ: ใส่ ID ของนักเรียนที่ต้องการเลื่อนไปเทอม 2'
        ws['A2'] = 'student_id'
        ws.append(['1', 'ตัวอย่าง: ID นักเรียนที่ต้องการเลื่อน'])
        headers = ['student_id']
        filename = 'promote_mid_term_template.xlsx'
    else:  # end_of_year
        ws['A1'] = 'คำแนะนำ: ใส่ ID ของนักเรียนและชั้นเรียนใหม่'
        ws['A2'] = 'student_id'
        ws['B2'] = 'new_grade_level'
        ws.append(['1', 'ป.2', 'ตัวอย่าง: ID และชั้นเรียนใหม่'])
        headers = ['student_id', 'new_grade_level']
        filename = 'promote_end_of_year_template.xlsx'
    
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    return StreamingResponse(stream, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)


@router.post("/promote_from_file")
def promote_students_from_file(
    file: UploadFile = File(...),
    promotion_type: str = None,
    new_grade_level: str = None,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    """Promote students from Excel file"""
    if getattr(current_user, 'role', None) != 'admin':
        raise HTTPException(status_code=403, detail='Only admins can promote students')
    
    if openpyxl is None:
        raise HTTPException(status_code=500, detail='Server missing openpyxl dependency')
    
    if promotion_type not in ['mid_term', 'end_of_year']:
        raise HTTPException(status_code=400, detail='Invalid promotion_type')
    
    if promotion_type == 'end_of_year' and not new_grade_level:
        raise HTTPException(status_code=400, detail='new_grade_level is required for end_of_year promotion')
    
    try:
        content = file.file.read()
        file.file.seek(0)
        
        # data_only=True reads calculated formula values instead of formula strings
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        
        promoted_count = 0
        failed_count = 0
        errors = []
        promoted_students = []
        
        # Get header row
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        idx = {col: i for i, col in enumerate(header_row)}
        
        # Process data rows (skip header)
        for r_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Get student_id
                student_id = None
                if 'student_id' in idx and row[idx['student_id']] is not None:
                    student_id = int(row[idx['student_id']])
                
                if not student_id:
                    failed_count += 1
                    errors.append({'row': r_i, 'error': 'Missing or invalid student_id'})
                    continue
                
                student = db.query(UserModel).filter(
                    UserModel.id == student_id,
                    UserModel.role == 'student'
                ).first()
                
                if not student:
                    failed_count += 1
                    errors.append({'row': r_i, 'error': f'Student ID {student_id} not found'})
                    continue
                
                old_grade = student.grade_level or 'ไม่ระบุ'
                
                if promotion_type == 'mid_term':
                    # Move to semester 2
                    if 'เทอม 2' not in str(student.grade_level or ''):
                        if 'เทอม 1' in str(student.grade_level or ''):
                            student.grade_level = str(student.grade_level).replace('เทอม 1', 'เทอม 2')
                        else:
                            student.grade_level = f"{student.grade_level} (เทอม 2)" if student.grade_level else "เทอม 2"
                
                elif promotion_type == 'end_of_year':
                    # Set new grade level
                    new_grade = new_grade_level
                    if 'new_grade_level' in idx and row[idx['new_grade_level']] is not None:
                        new_grade = str(row[idx['new_grade_level']]).strip()
                    
                    student.grade_level = new_grade
                
                db.add(student)
                db.flush()
                promoted_count += 1
                promoted_students.append({
                    'id': student.id,
                    'username': student.username,
                    'full_name': student.full_name,
                    'old_grade': old_grade,
                    'new_grade': student.grade_level
                })
                
            except Exception as e:
                failed_count += 1
                errors.append({'row': r_i, 'error': str(e)})
        
        db.commit()
        
        return {
            'success': True,
            'promotion_type': promotion_type,
            'promoted_count': promoted_count,
            'failed_count': failed_count,
            'promoted_students': promoted_students,
            'errors': errors,
            'message': f'เลื่อนชั้นเรียนสำเร็จ {promoted_count} นักเรียน, ล้มเหลว {failed_count} นักเรียน'
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f'Failed to process file: {str(e)}')